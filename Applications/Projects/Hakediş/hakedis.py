import time
import sys
from openpyxl import Workbook,load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

driver = webdriver.Chrome()
wait = WebDriverWait(driver, 15)
temp = ""
ilk = 0

def login_gspn(website):
    driver.get(website)
    Alert(driver).accept()
    username = wait.until(EC.presence_of_element_located((By.XPATH, """//*[@id="login_form_all"]/div[1]/dl/dd[1]/input""")))
    password = wait.until(EC.presence_of_element_located((By.XPATH, """//*[@id="login_form_all"]/div[1]/dl/dd[2]/input""")))
    usercr = get_username_password()
    username.send_keys(usercr[0])
    password.send_keys(usercr[1])
    login_form = wait.until(EC.presence_of_element_located((By.XPATH, """//*[@id="login_form_all"]/div[1]/img""")))
    login_form.click()
    Alert(driver).accept()

def go_management():
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"""/html/frameset/frame[2]""")))
    management = wait.until(EC.element_to_be_clickable((By.XPATH, """//*[@id="MAIN_04"]/span""")))
    management.click()

def go_work_order():
    global ilk
    driver.switch_to_default_content()	
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"""/html/frameset/frame[3]""")))	
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"""//*[@id="leftMenus"]""")))	
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"""//*[@id="b2BLeftMenuScroll"]""")))	
    # st = wait.until(EC.presence_of_element_located((By.XPATH, """/html/body/table/tbody/tr[2]/td/table/tbody/tr[1]/td"""))) #Favoriler	
    if(ilk==0):
        st = wait.until(EC.presence_of_element_located((By.XPATH, """/html/body/table/tbody/tr[10]/td""")))
        st.click()	
        ilk = 1
    ie = wait.until(EC.presence_of_element_located((By.XPATH, """/html/body/table/tbody/tr[11]/td/table/tbody/tr[6]/td""")))
    ie.click()

def get_name_surname(is_emri):
    driver.switch_to_default_content()
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"""/html/frameset/frame[3]""")))
    wait.until(EC.frame_to_be_available_and_switch_to_it((By.XPATH,"""//*[@id="rightContents"]""")))
    input_is_emri = wait.until(EC.presence_of_element_located((By.ID, """service_order_no""")))
    input_is_emri.send_keys(Keys.DELETE)
    input_is_emri.send_keys(is_emri)
    input_is_emri.send_keys(Keys.ENTER)
    name_surname = wait.until(EC.presence_of_element_located((By.XPATH, """/html/body/table/tbody/tr[1]/td[1]/form/table[11]/tbody/tr[2]/td[1]"""))).text

    if(name_surname == "" or name_surname is None):
        name_surname = "No Result"

    return name_surname

def read_data_from_txt(path):
    file = open(path, "r") 
    Lines = file.readlines() 
    for line in Lines: 
        go_work_order()
        time.sleep(4)
        # print(line.strip(),get_name_surname(line.strip()))
        export_results(line.strip(),get_name_surname(line.strip()))
        time.sleep(5)

def get_username_password():
    dosya = open("D:\\user2.txt","r",encoding="utf8")	
    satir = dosya.readline() 	
    print(satir)	
    return satir.split(",")	

def export_results(res_is_emri,res_name_surname):
    wb = load_workbook("output.xlsx")
    ws = wb.active
    # ws = wb.create_sheet("Result")
    ws.append((res_is_emri,res_name_surname))
    wb.save("output.xlsx")
    wb.close()

# login_gspn("https://gspn1.samsungcsportal.com/")
# go_management()
# read_data_from_txt('hakedisnote.txt')

